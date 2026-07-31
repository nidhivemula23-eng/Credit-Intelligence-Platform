"""
load_data.py loads a completed staging workbook (like auto_sector_staging.xlsx)
into your PostgreSQL 'credit_intel' database.

WHAT THIS SCRIPT DOES:
- Opens your Excel file
- For each of the 3 sheets (Income_Statement, Balance_Sheet, Cash_Flow):
    - Reads every row
    - Looks up the company's ID number from the Companies table (matching
      by company_name; this is why company names in your Excel must be
      spelled EXACTLY as they are in the Companies table)
    - Inserts that row into the matching SQL table
    - If a row for that company+year already exists, it UPDATES instead of creating a duplicate
"""

import re
import openpyxl
import psycopg2


# DATABASE CONNECTION

DB_CONFIG = {
    "host": "localhost",
    "port": 5432,
    "dbname": "credit_intel",
    "user": "postgres",
    "password": "YOUR PASSWORD HERE",   
}


# WHICH FILE TO LOAD 

EXCEL_FILE = "telecom_sector_staging.xlsx"


# Canonical column names per table (must match 01_schema.sql exactly)

SHEET_CONFIG = {
    "Income_Statement": {
        "table": "Income_Statement",
        "columns": ["revenue", "ebitda", "ebit", "interest_expense",
                    "depreciation", "tax", "net_profit"],
    },
    "Balance_Sheet": {
        "table": "Balance_Sheet",
        "columns": ["total_assets", "total_equity", "total_debt",
                    "current_assets", "current_liabilities",
                    "inventory", "cash_and_equivalents"],
    },
    "Cash_Flow": {
        "table": "Cash_Flow",
        "columns": ["cfo", "cfi", "cff", "capex"],
    },
}


def normalize_header(raw_header):
    """
    Turns messy headers like 'tax (pbt-eat)' or 'capex(cfo-fcf)' into
    clean canonical names like 'tax' and 'capex', by stripping anything
    in parentheses and extra whitespace. This means you can keep adding
    clarifying notes in parentheses in future sheets without breaking
    the loader.
    """
    cleaned = re.sub(r"\(.*?\)", "", str(raw_header))
    return cleaned.strip().lower().replace(" ", "_")


def get_company_id_map(cur):
    cur.execute("SELECT company_id, company_name FROM Companies;")
    return {name: cid for cid, name in cur.fetchall()}


def load_sheet(cur, ws, config, company_map):
    table = config["table"]
    expected_cols = config["columns"]

    # Read header row, normalize each header
    raw_headers = [cell.value for cell in ws[1]]
    normalized_headers = [normalize_header(h) for h in raw_headers]

    rows_loaded = 0
    rows_skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(normalized_headers, row))

        company_name = row_dict.get("company_name")
        fiscal_year = row_dict.get("fiscal_year")

        if company_name not in company_map:
            print(f"  SKIPPED — unknown company '{company_name}' "
                  f"(check spelling matches Companies table exactly)")
            rows_skipped += 1
            continue

        company_id = company_map[company_name]

        # Build the column list and values for this row, in schema order
        data_cols = ["company_id", "fiscal_year"] + expected_cols
        data_vals = [company_id, fiscal_year] + [
            row_dict.get(col) for col in expected_cols
        ]

        placeholders = ", ".join(["%s"] * len(data_vals))
        col_names = ", ".join(data_cols)
        update_clause = ", ".join(
            f"{col} = EXCLUDED.{col}" for col in expected_cols
        )

        sql = f"""
            INSERT INTO {table} ({col_names})
            VALUES ({placeholders})
            ON CONFLICT (company_id, fiscal_year)
            DO UPDATE SET {update_clause};
        """
        cur.execute(sql, data_vals)
        rows_loaded += 1

    print(f"  {table}: {rows_loaded} rows loaded, {rows_skipped} skipped")


def main():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    company_map = get_company_id_map(cur)
    print(f"Found {len(company_map)} companies in database.\n")

    for sheet_name, config in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found in {EXCEL_FILE}, skipping.")
            continue
        print(f"Loading {sheet_name}...")
        load_sheet(cur, wb[sheet_name], config, company_map)

    conn.commit()
    cur.close()
    conn.close()
    print("\nDone. All changes committed to credit_intel database.")


if __name__ == "__main__":
    main()
